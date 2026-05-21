#!/usr/bin/env python3
"""
Export a GanttBok job into a pre-styled XLSX template, preserving the template's
weekly grid (Mon-Sun columns, merged Week-N labels, weekend shading) and replacing
the dates so col D = Monday of the project start week.

This was the prototype for what might become an in-app "Export to Excel template"
feature in GanttBok.

Usage:
    python3 export_to_xlsx_template.py \\
        --template "/path/to/PROJECT TIMELINE TEMPLATE.xlsx" \\
        --db ~/Library/Application\\ Support/Gantt\\ Bok/ganttbok.db \\
        --job-id 5 \\
        --output "/path/to/output.xlsx"

Or edit the defaults below and just run `python3 export_to_xlsx_template.py`.

How it works
------------
1. Read job + phases + tasks + no_work_days from the SQLite DB.
2. Open the template XLSX and walk every existing date cell in rows 1-3,
   shifting it by N days so col D becomes the Monday of the project start week.
   Row 4 (M/T/W/T/F/S/S day-letters) stays correct because we only shift by whole
   weeks; weekend shading is preserved cell-by-cell. Row-1 ISO week labels are
   recomputed.
3. Wipe values from row 6 onwards (keeps cell formatting/weekend shading intact).
4. Write project name at A6, then for each phase: phase name in col B (bold),
   each task name in col C, and a coloured fill from start-col to end-col on
   each task row. Bars naturally span calendar days (weekends included) and
   stretch by one day per SA public holiday they cross — same logic GanttBok
   uses internally.

Limitations
-----------
- The template's date range is finite (26 weeks). Projects extending beyond the
  template's last column get clamped with a warning. To extend, modify the
  template or use the longer-running version of this script in earlier git
  history.
- Templates with embedded charts, images, or complex conditional formatting may
  need extra handling — openpyxl preserves most of this but pivot tables and
  defined names are sometimes stripped.
"""

import argparse
from copy import copy
from datetime import date, datetime, timedelta
import sqlite3
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Need openpyxl: pip install openpyxl")


DEFAULTS = {
    "template": "/Users/cncuser/Desktop/NOORDHOEK_DATA/Project_managment/PROJECT TIMELINE TEMPLATE.xlsx",
    "db": "/Users/cncuser/Library/Application Support/Gantt Bok/ganttbok.db",
    "job_id": 5,
    "output": "/Users/cncuser/Desktop/NOORDHOEK_DATA/Project_managment/PROJECT TIMELINE - Noordhoek House.xlsx",
    # Where the date columns begin in the template. Col D = 4.
    "col_offset": 4,
    # Where the first project row goes (matches template's row 6 = first project name).
    "first_project_row": 6,
}


def load_job(db_path, job_id):
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    job = dict(cur.execute("SELECT * FROM job WHERE id=?", (job_id,)).fetchone())
    phases = [
        dict(p)
        for p in cur.execute(
            "SELECT * FROM phase WHERE job_id=? ORDER BY order_index",
            (job_id,),
        ).fetchall()
    ]
    tasks_by_phase = {
        p["id"]: [
            dict(t)
            for t in cur.execute(
                "SELECT * FROM task WHERE phase_id=? ORDER BY order_index",
                (p["id"],),
            ).fetchall()
        ]
        for p in phases
    }
    no_work = {
        row[0]
        for row in cur.execute(
            "SELECT date FROM no_work_day WHERE job_id=?", (job_id,)
        ).fetchall()
    }
    return job, phases, tasks_by_phase, no_work


def snap_to_workday(d, no_work_set):
    while d.weekday() >= 5 or d.isoformat() in no_work_set:
        d += timedelta(days=1)
    return d


def task_span(start_iso, duration_workdays, no_work_set):
    """Return (start_date, end_date) inclusive — calendar span the bar should fill."""
    d = datetime.strptime(start_iso, "%Y-%m-%d").date()
    d = snap_to_workday(d, no_work_set)
    start_d = d
    count = 1
    while count < duration_workdays:
        d += timedelta(days=1)
        if d.weekday() < 5 and d.isoformat() not in no_work_set:
            count += 1
    return start_d, d


def shift_template_dates(ws, new_cal_start, col_offset, max_col):
    """Shift every date in rows 1-3 so col D == new_cal_start (a Monday)."""
    old_cal_start = ws.cell(3, col_offset).value.date()
    shift = new_cal_start - old_cal_start
    for col in range(col_offset, max_col + 1):
        # Row 3: every-column date
        v = ws.cell(3, col).value
        if isinstance(v, datetime):
            nd = v.date() + shift
            ws.cell(3, col).value = datetime(nd.year, nd.month, nd.day)
        # Row 2: Monday date (sparse)
        v = ws.cell(2, col).value
        if isinstance(v, datetime):
            nd = v.date() + shift
            ws.cell(2, col).value = datetime(nd.year, nd.month, nd.day)
        # Row 1: Week label — recompute ISO week from the shifted Monday
        v = ws.cell(1, col).value
        if isinstance(v, str) and v.startswith("Week "):
            r3 = ws.cell(3, col).value
            if isinstance(r3, datetime):
                ws.cell(1, col).value = f"Week {r3.date().isocalendar().week}"


def wipe_project_rows(ws, first_row, max_col, weekday_fill, weekend_fill, col_offset):
    """Clear values AND reset fills to template defaults from first_row downward.
    Weekday columns get the empty/no-fill style; Sat/Sun columns get the gray shade.
    This ensures old phase-title rows (which had a blue tint across the whole row)
    revert to a clean grid before we re-apply the blue fill to OUR phase rows.
    """
    for r in range(first_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            if c < col_offset:
                # Cols A/B/C — narrow label cols, always clear
                cell.fill = copy(weekday_fill)
            else:
                # Determine weekday from row 3 (post-shift) for this column
                day_cell = ws.cell(3, c).value
                if isinstance(day_cell, datetime) and day_cell.weekday() >= 5:
                    cell.fill = copy(weekend_fill)
                else:
                    cell.fill = copy(weekday_fill)


def write_job(ws, job, phases, tasks_by_phase, no_work, col_offset, max_col, first_row, new_cal_start, phase_header_fill=None, phase_header_font=None):
    def col_for(d):
        return col_offset + (d - new_cal_start).days

    row = first_row
    ws.cell(row, 1).value = job["name"]
    ws.cell(row, 1).font = Font(bold=True, size=10)
    row += 1

    bar_count = 0
    for p in phases:
        if not tasks_by_phase[p["id"]]:
            continue
        # Phase header is a banner row: fill the whole row in the template's blue tint.
        if phase_header_fill is not None:
            for c in range(1, max_col + 1):
                ws.cell(row, c).fill = copy(phase_header_fill)
        phase_cell = ws.cell(row, 2)
        phase_cell.value = p["name"]
        phase_cell.font = copy(phase_header_font) if phase_header_font is not None else Font(bold=True, size=10)
        row += 1
        colour_hex = p["colour"].lstrip("#")
        for t in tasks_by_phase[p["id"]]:
            ws.cell(row, 3).value = t["name"]
            ws.cell(row, 3).font = Font(size=10)
            s, e = task_span(t["start_date"], t["duration_workdays"], no_work)
            c0 = col_for(s)
            c1 = col_for(e)
            if c0 < col_offset or c1 > max_col:
                print(
                    f'  WARNING: "{t["name"]}" {s}..{e} (cols {c0}..{c1}) outside template range; clamping'
                )
                c0 = max(c0, col_offset)
                c1 = min(c1, max_col)
            bar = PatternFill(
                start_color=f"FF{colour_hex}",
                end_color=f"FF{colour_hex}",
                fill_type="solid",
            )
            for c in range(c0, c1 + 1):
                ws.cell(row, c).fill = copy(bar)
            bar_count += 1
            row += 1
        row += 1  # blank between phases
    return bar_count, row


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template", default=DEFAULTS["template"])
    ap.add_argument("--db", default=DEFAULTS["db"])
    ap.add_argument("--job-id", type=int, default=DEFAULTS["job_id"])
    ap.add_argument("--output", default=DEFAULTS["output"])
    args = ap.parse_args()

    job, phases, tasks_by_phase, no_work = load_job(args.db, args.job_id)
    print(f'Job: "{job["name"]}" — {len(phases)} phases, '
          f'{sum(len(v) for v in tasks_by_phase.values())} tasks, '
          f'{len(no_work)} no-work days')

    wb = openpyxl.load_workbook(args.template)
    ws = wb.active
    max_col = ws.max_column

    # Capture the template's reference styles BEFORE wiping anything.
    # - Phase header: B7 ("Coffee shop") — full-row blue tint
    # - Weekday default: E8 (task row, Tue) — no fill
    # - Weekend default: I8 (task row, Sat) — gray shade
    phase_header_fill = copy(ws["B7"].fill)
    phase_header_font = copy(ws["B7"].font)
    weekday_default_fill = copy(ws["E8"].fill)
    weekend_default_fill = copy(ws["I8"].fill)

    proj_start = datetime.strptime(job["project_start_date"], "%Y-%m-%d").date()
    new_cal_start = proj_start - timedelta(days=proj_start.weekday())
    print(f"Project start {proj_start} -> calendar starts {new_cal_start}")

    shift_template_dates(ws, new_cal_start, DEFAULTS["col_offset"], max_col)
    wipe_project_rows(
        ws, DEFAULTS["first_project_row"], max_col,
        weekday_default_fill, weekend_default_fill, DEFAULTS["col_offset"],
    )
    bars, end_row = write_job(
        ws, job, phases, tasks_by_phase, no_work,
        DEFAULTS["col_offset"], max_col,
        DEFAULTS["first_project_row"], new_cal_start,
        phase_header_fill=phase_header_fill,
        phase_header_font=phase_header_font,
    )
    wb.save(args.output)
    print(f"OK — {bars} bars, ends at row {end_row}, saved to {args.output}")


if __name__ == "__main__":
    main()
